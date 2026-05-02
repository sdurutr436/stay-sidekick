"""Servicio del mapa de calor.

Orquesta la generación de datos de calor desde PMS (Smoobu) o desde XLSX.
Los datos procesados NO se persisten (cumplimiento RGPD): todo en memoria.
"""

from __future__ import annotations

import io
import logging
from datetime import date, timedelta

import openpyxl
import pandas as pd
import requests
from marshmallow import ValidationError

from app.common.crypto import decrypt
from app.extensions import db
from app.h_mapa_de_calor import repository as repo
from app.h_mapa_de_calor.schemas import ConfigXlsxSchema, UmbralesSchema
from app.h_maestro_apartamentos import repository as apt_repo
from app.normalizador_pms.factory import build_pms_client
from app.h_sincronizador_contactos.model import ESTADO_EXITO

logger = logging.getLogger(__name__)

ORIGEN_HEATMAP_PMS = "heatmap_pms"
ORIGEN_HEATMAP_XLSX = "heatmap_xlsx"

_umbrales_schema = UmbralesSchema()
_config_xlsx_schema = ConfigXlsxSchema()


# ── Umbrales ──────────────────────────────────────────────────────────────


def get_umbrales(empresa_id: str) -> dict:
    return repo.get_umbrales(empresa_id)


def save_umbrales(empresa_id: str, json_data: dict) -> tuple[dict | None, list[str]]:
    try:
        clean = _umbrales_schema.load(json_data)
    except ValidationError as exc:
        return None, _flatten(exc.messages)
    result = repo.save_umbrales(empresa_id, clean)
    db.session.commit()
    return result, []


# ── Config XLSX ───────────────────────────────────────────────────────────


def get_config_xlsx(empresa_id: str) -> dict:
    return repo.get_config_xlsx(empresa_id)


def save_config_xlsx(empresa_id: str, json_data: dict) -> tuple[dict | None, list[str]]:
    try:
        clean = _config_xlsx_schema.load(json_data)
    except ValidationError as exc:
        return None, _flatten(exc.messages)
    result = repo.save_config_xlsx(
        empresa_id,
        col_fecha_checkin=clean["col_fecha_checkin"],
        col_fecha_checkout=clean.get("col_fecha_checkout"),
    )
    db.session.commit()
    return result, []


# ── Generación desde PMS ──────────────────────────────────────────────────


def generar_desde_pms(
    empresa_id: str,
    desde: date,
    hasta: date,
) -> tuple[list[dict] | None, str | None]:
    """Genera los datos del mapa de calor llamando al PMS configurado."""
    pms_config = apt_repo.get_pms_config(empresa_id)
    if not pms_config or not pms_config.api_key_cifrada:
        return None, "No hay PMS conectado. Usa el endpoint XLSX."

    api_key = decrypt(pms_config.api_key_cifrada)
    if api_key is None:
        return None, "No se pudo descifrar la API key del PMS."

    desde_str = desde.isoformat()
    hasta_str = hasta.isoformat()

    try:
        client = build_pms_client(pms_config.proveedor, api_key, pms_config.endpoint)
        reservas_checkin = client.fetch_reservations(desde=desde_str, hasta=hasta_str)
        reservas_checkout = client.fetch_by_departure(desde_str, hasta_str)
    except (requests.RequestException, ValueError, NotImplementedError) as exc:
        logger.error("Error al obtener reservas del PMS para heatmap: %s", exc)
        return None, f"Error al obtener datos del PMS: {exc}"

    checkins: dict[str, int] = {}
    for r in reservas_checkin:
        if r.checkin:
            checkins[r.checkin] = checkins.get(r.checkin, 0) + 1

    checkouts: dict[str, int] = {}
    for r in reservas_checkout:
        if r.checkout:
            checkouts[r.checkout] = checkouts.get(r.checkout, 0) + 1

    dias = _construir_grid(desde, hasta, checkins, checkouts)
    num_con_datos = sum(1 for d in dias if d["checkins"] > 0 or d["checkouts"] > 0)

    repo.create_sync_log(
        empresa_id=empresa_id,
        origen=ORIGEN_HEATMAP_PMS,
        estado=ESTADO_EXITO,
        num_registros=num_con_datos,
        detalle=f"Rango: {desde_str} — {hasta_str}",
    )
    db.session.commit()

    return dias, None


# ── Generación desde XLSX ─────────────────────────────────────────────────


def generar_desde_xlsx(
    empresa_id: str,
    checkins_bytes: bytes,
    checkouts_bytes: bytes | None,
    desde: date,
    hasta: date,
) -> tuple[list[dict] | None, list[str], str | None]:
    """Genera datos del mapa de calor a partir de ficheros XLSX subidos.

    Returns (dias, warnings, error). Si error no es None, dias es None.
    """
    cfg = repo.get_config_xlsx(empresa_id)
    col_checkin = cfg.get("col_fecha_checkin")
    col_checkout = cfg.get("col_fecha_checkout")

    if not col_checkin:
        return None, [], (
            "Configura la columna de fecha de check-in en el perfil de empresa "
            "antes de usar esta herramienta."
        )

    if checkouts_bytes is not None and not col_checkout:
        return None, [], (
            "Configura la columna de fecha de check-out en el perfil de empresa "
            "antes de usar el fichero de salidas."
        )

    warnings: list[str] = []

    checkins, w_ci, error = _parsear_xlsx_fechas(checkins_bytes, col_checkin)
    if error:
        return None, [], error
    warnings.extend(w_ci)

    checkouts: dict[str, int] = {}
    if checkouts_bytes is not None:
        checkouts, w_co, error = _parsear_xlsx_fechas(checkouts_bytes, col_checkout)
        if error:
            return None, [], error
        warnings.extend(w_co)

    dias = _construir_grid(desde, hasta, checkins, checkouts)
    num_con_datos = sum(1 for d in dias if d["checkins"] > 0 or d["checkouts"] > 0)

    repo.create_sync_log(
        empresa_id=empresa_id,
        origen=ORIGEN_HEATMAP_XLSX,
        estado=ESTADO_EXITO,
        num_registros=num_con_datos,
        detalle=f"Rango: {desde.isoformat()} — {hasta.isoformat()}",
    )
    db.session.commit()

    return dias, warnings, None


# ── Helpers internos ──────────────────────────────────────────────────────


def _sanitize_xlsx(file_bytes: bytes) -> openpyxl.Workbook:
    """Carga el XLSX y elimina hojas ocultas y celdas con fórmulas/inyección."""
    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes),
        read_only=False,
        data_only=True,
    )
    for name in wb.sheetnames[:]:
        if wb[name].sheet_state != "visible":
            del wb[name]
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value and cell.value[0] in ("=", "+", "-", "@"):
                    cell.value = ""
    return wb


def _parsear_xlsx_fechas(
    file_bytes: bytes,
    col_nombre: str,
) -> tuple[dict[str, int], list[str], str | None]:
    """Extrae y agrega fechas de la columna indicada en el XLSX.

    Devuelve ({fecha_iso: count}, warnings, error). Si error no es None el resto es vacío.
    """
    try:
        wb = _sanitize_xlsx(file_bytes)
    except Exception as exc:
        return {}, [], f"No se pudo abrir el archivo Excel: {exc}"

    try:
        ws = wb.active
        if ws is None:
            return {}, [], "El archivo Excel no tiene hojas de cálculo."

        data = list(ws.values)
    finally:
        wb.close()

    if len(data) < 2:
        return {}, [], "El archivo debe tener cabeceras en la primera fila y al menos una fila de datos."

    headers = [str(h).strip() if h is not None else "" for h in data[0]]
    col_lower = col_nombre.strip().lower()
    idx = next(
        (i for i, h in enumerate(headers) if h.strip().lower() == col_lower),
        None,
    )
    if idx is None:
        return {}, [], (
            f"Columna '{col_nombre}' no encontrada en el archivo. "
            "Comprueba el nombre configurado en el perfil de empresa."
        )

    df = pd.DataFrame(data[1:], columns=headers)
    serie = df.iloc[:, idx]

    conteo: dict[str, int] = {}
    descartadas = 0

    for val in serie:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            descartadas += 1
            continue
        try:
            d = pd.to_datetime(val, dayfirst=True, errors="raise").date()
            iso = d.isoformat()
            conteo[iso] = conteo.get(iso, 0) + 1
        except Exception:
            descartadas += 1

    warnings: list[str] = []
    if descartadas:
        warnings.append(f"{descartadas} filas descartadas por fecha inválida")

    return conteo, warnings, None


def _construir_grid(
    desde: date,
    hasta: date,
    checkins: dict[str, int],
    checkouts: dict[str, int],
) -> list[dict]:
    """Genera la lista de días del rango, completando semanas completas (lun–dom).

    Los días fuera del rango [desde, hasta] quedan marcados con mes_adyacente=True.
    """
    # Expandir a semanas completas
    inicio = desde - timedelta(days=desde.weekday())  # lunes de la primera semana
    fin = hasta + timedelta(days=6 - hasta.weekday())  # domingo de la última semana

    dias: list[dict] = []
    dia = inicio
    while dia <= fin:
        iso = dia.isoformat()
        dias.append({
            "fecha": iso,
            "checkins": checkins.get(iso, 0),
            "checkouts": checkouts.get(iso, 0),
            "mesAdyacente": dia < desde or dia > hasta,
        })
        dia += timedelta(days=1)

    return dias


def _flatten(messages: dict) -> list[str]:
    errors: list[str] = []
    for field, msgs in messages.items():
        if isinstance(msgs, list):
            for msg in msgs:
                errors.append(f"{field}: {msg}")
        elif isinstance(msgs, str):
            errors.append(f"{field}: {msgs}")
        elif isinstance(msgs, dict):
            for sub in msgs.values():
                if isinstance(sub, list):
                    errors.extend(f"{field}: {m}" for m in sub)
    return errors
