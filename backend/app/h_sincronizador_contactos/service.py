"""Servicio de sincronización con Google Contacts.

Orquesta el flujo completo:
1. OAuth 2.0: generación de URL, intercambio de código, almacenamiento de tokens.
2. Sincronización: PMS/XLSX → agrupar por nombre-teléfono → Google People API.
3. Fallback: exportación CSV para importación manual.

Los datos de huéspedes se procesan en memoria y NO se persisten en BD
(cumplimiento RGPD).
"""

import logging
import re
import secrets
from io import BytesIO
from urllib.parse import urlencode

import requests
from flask import current_app
from marshmallow import ValidationError
from openpyxl import load_workbook

from app.common.crypto import decrypt, encrypt
from app.extensions import db
from app.h_sincronizador_contactos.model import (
    IntegracionGoogle,
    ORIGEN_GOOGLE_CONTACTS,
    ESTADO_EXITO,
    ESTADO_ERROR,
    ESTADO_PARCIAL,
)
from app.h_sincronizador_contactos import repository as repo
from app.h_sincronizador_contactos.schemas import PreferenciasContactosSchema, SyncRangoSchema
from app.h_sincronizador_contactos.csv_export import build_csv
from app.h_sincronizador_contactos.google_people_client import GooglePeopleClient
from app.h_sincronizador_contactos.contacto_formatter import (
    ContactoAgrupado,
    parsear_fecha,
    parsear_telefono,
    agrupar_contactos,
    aplicar_plantilla,
)
from app.normalizador_pms.factory import build_pms_client
from app.h_maestro_apartamentos import repository as apt_repo

logger = logging.getLogger(__name__)

_GOOGLE_AUTH_URL = "https://accounts.google.com/o/oauth2/v2/auth"
_GOOGLE_TOKEN_URL = "https://oauth2.googleapis.com/token"
_SCOPES = "https://www.googleapis.com/auth/contacts"

# Regex para extraer IDs de tipología del formato {12345} o {12345}{67890}
_RE_IDS_TIP = re.compile(r"\{(\d+)\}")

_pref_schema = PreferenciasContactosSchema()
_sync_schema = SyncRangoSchema()


# ── OAuth 2.0 ────────────────────────────────────────────────────────────


def build_oauth_url(empresa_id: str) -> tuple[str, str]:
    """Genera la URL de autorización de Google OAuth 2.0."""
    client_id = current_app.config["GOOGLE_CLIENT_ID"]
    redirect_uri = current_app.config["GOOGLE_REDIRECT_URI"]
    state = secrets.token_urlsafe(32)

    params = {
        "client_id": client_id,
        "redirect_uri": redirect_uri,
        "response_type": "code",
        "scope": _SCOPES,
        "access_type": "offline",
        "prompt": "consent",
        "state": state,
    }
    query = urlencode(params)
    return f"{_GOOGLE_AUTH_URL}?{query}", state


def exchange_code_for_tokens(code: str, empresa_id: str) -> tuple[bool, str | None]:
    """Intercambia el código OAuth por tokens y los persiste cifrados."""
    client_id = current_app.config["GOOGLE_CLIENT_ID"]
    client_secret = current_app.config["GOOGLE_CLIENT_SECRET"]
    redirect_uri = current_app.config["GOOGLE_REDIRECT_URI"]

    try:
        resp = requests.post(
            _GOOGLE_TOKEN_URL,
            data={
                "code": code,
                "client_id": client_id,
                "client_secret": client_secret,
                "redirect_uri": redirect_uri,
                "grant_type": "authorization_code",
            },
            timeout=20,
        )
        resp.raise_for_status()
        data = resp.json()
    except requests.RequestException as exc:
        logger.error("Error al intercambiar código OAuth: %s", exc)
        return False, f"Error al conectar con Google: {exc}"

    refresh_token = data.get("refresh_token")
    if not refresh_token:
        return False, (
            "Google no devolvió refresh_token. Revoca el acceso desde tu cuenta Google "
            "y vuelve a conectar."
        )

    access_token = data.get("access_token", "")
    expires_in = data.get("expires_in", 3600)
    alcance = data.get("scope", _SCOPES)

    from datetime import datetime, timedelta, timezone
    token_expiry = datetime.now(timezone.utc) + timedelta(seconds=expires_in)

    repo.upsert_google_tokens(
        empresa_id=empresa_id,
        access_token_cifrado=encrypt(access_token),
        refresh_token_cifrado=encrypt(refresh_token),
        token_expiry=token_expiry,
        alcance=alcance,
    )
    db.session.commit()
    return True, None


def get_google_status(empresa_id: str) -> dict:
    """Devuelve el estado de la integración Google de la empresa."""
    integracion = repo.get_google_integration(empresa_id)
    if not integracion or not integracion.activo:
        return {"conectado": False}
    return {
        "conectado": True,
        "alcance": integracion.alcance,
        "ultimo_sync": _ultimo_sync(empresa_id),
    }


def disconnect_google(empresa_id: str) -> str | None:
    """Elimina los tokens Google de la empresa. Devuelve error o None."""
    integracion = repo.get_google_integration(empresa_id)
    if not integracion:
        return "No hay cuenta Google conectada."
    repo.delete_google_integration(integracion)
    db.session.commit()
    return None


# ── Preferencias ─────────────────────────────────────────────────────────


def get_preferencias(empresa_id: str) -> dict:
    return repo.get_preferencias_contactos(empresa_id)


def save_preferencias(empresa_id: str, json_data: dict) -> tuple[dict | None, list[str]]:
    try:
        clean = _pref_schema.load(json_data)
    except ValidationError as exc:
        return None, _flatten(exc.messages)

    result = repo.save_preferencias_contactos(empresa_id, clean)
    db.session.commit()
    return result, []


# ── Sincronización ────────────────────────────────────────────────────────


def sync_contacts(empresa_id: str, json_data: dict) -> tuple[dict | None, str | None]:
    """Sincroniza contactos desde el PMS hacia Google Contacts."""
    integracion = repo.get_google_integration(empresa_id)
    if not integracion or not integracion.activo:
        return None, "No hay cuenta Google conectada. Conéctala primero desde el panel."

    pms_config = apt_repo.get_pms_config(empresa_id)
    if not pms_config or not pms_config.api_key_cifrada:
        return None, "No hay configuración de PMS. Configura tu API key primero."

    api_key_pms = decrypt(pms_config.api_key_cifrada)
    if api_key_pms is None:
        return None, "No se pudo descifrar la API key del PMS."

    refresh_token = decrypt(integracion.refresh_token_cifrado)
    access_token = (
        decrypt(integracion.access_token_cifrado)
        if integracion.access_token_cifrado
        else None
    )
    if refresh_token is None:
        return None, "No se pudo descifrar el refresh_token de Google."

    try:
        pms_client = build_pms_client(pms_config.proveedor, api_key_pms, pms_config.endpoint)
        desde_str = json_data.get("desde")
        hasta_str = json_data.get("hasta")
        reservas = pms_client.fetch_reservations(
            desde=desde_str.isoformat() if desde_str else None,
            hasta=hasta_str.isoformat() if hasta_str else None,
        )
    except (requests.RequestException, NotImplementedError) as exc:
        logger.error("Error al obtener reservas del PMS: %s", exc)
        _log_sync(empresa_id, ESTADO_ERROR, 0, f"Error PMS: {exc}")
        db.session.commit()
        return None, f"Error al obtener reservas del PMS: {exc}"

    prefs = repo.get_preferencias_contactos(empresa_id)
    contactos = _reservas_a_contactos(reservas)

    client_id = current_app.config["GOOGLE_CLIENT_ID"]
    client_secret = current_app.config["GOOGLE_CLIENT_SECRET"]

    google_client = GooglePeopleClient(
        client_id=client_id,
        client_secret=client_secret,
        access_token=access_token,
        refresh_token=refresh_token,
        token_expiry=integracion.token_expiry,
    )

    nuevos = 0
    actualizados = 0
    errores: list[str] = []

    for contacto in contactos:
        try:
            payload = _build_contact_payload(contacto, prefs)
            _, es_nuevo = google_client.upsert_contact(payload, contacto.telefono)
            if es_nuevo:
                nuevos += 1
            else:
                actualizados += 1
        except requests.RequestException as exc:
            logger.warning("Error al subir contacto '%s': %s", contacto.nombre, exc)
            errores.append(f"{contacto.nombre}: {exc}")

    if google_client.token_refreshed:
        new_at = google_client.current_access_token
        new_expiry = google_client.current_token_expiry
        if new_at:
            repo.update_access_token(integracion, encrypt(new_at), new_expiry)

    estado = ESTADO_PARCIAL if errores else ESTADO_EXITO
    _log_sync(
        empresa_id, estado, nuevos + actualizados,
        f"Nuevos: {nuevos}, actualizados: {actualizados}"
        + (f", errores: {len(errores)}" if errores else ""),
    )
    db.session.commit()

    result: dict = {
        "total": len(reservas),
        "nuevos": nuevos,
        "actualizados": actualizados,
    }
    if errores:
        result["advertencias"] = errores
    return result, None


# ── Export CSV ────────────────────────────────────────────────────────────


def export_csv(empresa_id: str, json_data: dict) -> tuple[bytes | None, str | None]:
    """Genera el CSV de contactos en formato Google para importación manual."""
    pms_config = apt_repo.get_pms_config(empresa_id)
    if not pms_config or not pms_config.api_key_cifrada:
        return None, "No hay configuración de PMS."

    api_key_pms = decrypt(pms_config.api_key_cifrada)
    if api_key_pms is None:
        return None, "No se pudo descifrar la API key del PMS."

    try:
        pms_client = build_pms_client(pms_config.proveedor, api_key_pms, pms_config.endpoint)
        desde_str = json_data.get("desde")
        hasta_str = json_data.get("hasta")
        reservas = pms_client.fetch_reservations(
            desde=desde_str.isoformat() if desde_str else None,
            hasta=hasta_str.isoformat() if hasta_str else None,
        )
    except (requests.RequestException, NotImplementedError) as exc:
        return None, f"Error al obtener reservas del PMS: {exc}"

    prefs = repo.get_preferencias_contactos(empresa_id)
    contactos = _reservas_a_contactos(reservas)
    csv_bytes = build_csv(contactos, prefs)
    return csv_bytes, None


# ── XLSX ─────────────────────────────────────────────────────────────────


def sync_from_xlsx(empresa_id: str, file_bytes: bytes) -> tuple[dict | None, str | None]:
    """Sincroniza contactos con Google People API desde un XLSX subido."""
    integracion = repo.get_google_integration(empresa_id)
    if not integracion or not integracion.activo:
        return None, "No hay cuenta Google conectada. Conéctala primero desde el panel."

    refresh_token = decrypt(integracion.refresh_token_cifrado)
    access_token = (
        decrypt(integracion.access_token_cifrado)
        if integracion.access_token_cifrado
        else None
    )
    if refresh_token is None:
        return None, "No se pudo descifrar el refresh_token de Google."

    prefs = repo.get_preferencias_contactos(empresa_id)
    registros, parse_errors = _parse_xlsx_contactos(file_bytes, empresa_id, prefs)
    if not registros and parse_errors:
        return None, parse_errors[0]

    contactos = agrupar_contactos(registros)
    client_id = current_app.config["GOOGLE_CLIENT_ID"]
    client_secret = current_app.config["GOOGLE_CLIENT_SECRET"]

    google_client = GooglePeopleClient(
        client_id=client_id,
        client_secret=client_secret,
        access_token=access_token,
        refresh_token=refresh_token,
        token_expiry=integracion.token_expiry,
    )

    nuevos = 0
    actualizados = 0
    errores: list[str] = list(parse_errors)

    for contacto in contactos:
        try:
            payload = _build_contact_payload(contacto, prefs)
            _, es_nuevo = google_client.upsert_contact(payload, contacto.telefono)
            if es_nuevo:
                nuevos += 1
            else:
                actualizados += 1
        except requests.RequestException as exc:
            logger.warning("Error al subir contacto '%s': %s", contacto.nombre, exc)
            errores.append(f"{contacto.nombre}: {exc}")

    if google_client.token_refreshed:
        new_at = google_client.current_access_token
        new_expiry = google_client.current_token_expiry
        if new_at:
            repo.update_access_token(integracion, encrypt(new_at), new_expiry)

    estado = ESTADO_PARCIAL if errores else ESTADO_EXITO
    _log_sync(
        empresa_id, estado, nuevos + actualizados,
        f"XLSX — Nuevos: {nuevos}, actualizados: {actualizados}"
        + (f", errores: {len(errores)}" if errores else ""),
    )
    db.session.commit()

    result: dict = {
        "total": len(contactos),
        "nuevos": nuevos,
        "actualizados": actualizados,
    }
    if errores:
        result["advertencias"] = errores
    return result, None


def export_csv_from_xlsx(empresa_id: str, file_bytes: bytes) -> tuple[bytes | None, str | None]:
    """Genera el CSV de Google Contacts desde un XLSX subido."""
    prefs = repo.get_preferencias_contactos(empresa_id)
    registros, parse_errors = _parse_xlsx_contactos(file_bytes, empresa_id, prefs)
    if not registros and parse_errors:
        return None, parse_errors[0]

    contactos = agrupar_contactos(registros)
    csv_bytes = build_csv(contactos, prefs)
    return csv_bytes, None


# ── Helpers ───────────────────────────────────────────────────────────────


def _reservas_a_contactos(reservas: list) -> list[ContactoAgrupado]:
    """Convierte lista de ReservaEstandar a ContactoAgrupado (filtra sin teléfono)."""
    registros = []
    for r in reservas:
        tel = parsear_telefono(r.telefono)
        if tel is None:
            continue  # skip silencioso: sin teléfono no hay contacto útil
        registros.append({
            "nombre": r.nombre_raw,
            "telefono": tel,
            "checkin_date": parsear_fecha(r.checkin),
            "apartamentos": [r.nombre_apartamento] if r.nombre_apartamento else [],
        })
    return agrupar_contactos(registros)


def _parse_xlsx_contactos(
    file_bytes: bytes,
    empresa_id: str,
    prefs: dict,
) -> tuple[list[dict], list[str]]:
    """Parsea XLSX de reservas con posiciones de columna configuradas.

    Usa los índices de columna de prefs["xlsx_reservas"] (1-indexado; 0 = detección
    por cabecera). Extrae IDs de tipología en formato {12345} y los cruza con el
    maestro de apartamentos. Sin teléfono → skip silencioso.

    Returns list de dicts con: nombre, telefono, checkin_date, apartamentos.
    """
    cols = prefs.get("xlsx_reservas") or {}
    col_checkin   = int(cols.get("col_checkin",   0))
    col_nombre    = int(cols.get("col_nombre",    0))
    col_tipologia = int(cols.get("col_tipologia", 0))
    col_telefono  = int(cols.get("col_telefono",  0))

    errors: list[str] = []

    try:
        wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        return [], [f"No se pudo abrir el archivo Excel: {exc}"]

    try:
        ws = wb.active
        if ws is None:
            return [], ["El archivo Excel no tiene hojas de cálculo."]

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return [], ["El archivo debe tener cabeceras en la primera fila y al menos una fila de datos."]

        headers = [
            str(h).strip().lower().replace(" ", "_") if h is not None else ""
            for h in rows[0]
        ]

        _HEADER_MAP = {
            "checkin":   ["checkin", "check-in", "check_in", "fecha_entrada", "arrival", "entrada"],
            "nombre":    ["nombre", "name", "guest", "huesped", "huésped", "guest_name", "referencia"],
            "tipologia": ["tipologia", "tipología", "id_tipologia", "unit_type", "tipo"],
            "telefono":  ["telefono", "teléfono", "phone", "tel", "movil", "móvil"],
        }

        def _find_col(configured: int, candidates: list[str]) -> int | None:
            if configured > 0:
                return configured - 1  # 1-indexado → 0-indexado
            for i, h in enumerate(headers):
                if h in candidates:
                    return i
            return None

        idx_checkin   = _find_col(col_checkin,   _HEADER_MAP["checkin"])
        idx_nombre    = _find_col(col_nombre,    _HEADER_MAP["nombre"])
        idx_tipologia = _find_col(col_tipologia, _HEADER_MAP["tipologia"])
        idx_telefono  = _find_col(col_telefono,  _HEADER_MAP["telefono"])

        if idx_nombre is None:
            return [], [
                "No se encontró columna de nombre del huésped. "
                "Configura la posición de columna en el perfil o añade una cabecera reconocida."
            ]

        registros: list[dict] = []
        for row in rows[1:]:
            def _cell(idx: int | None):
                return row[idx] if idx is not None and idx < len(row) else None

            nombre_raw = _cell(idx_nombre)
            if not nombre_raw:
                continue
            nombre = str(nombre_raw).strip()
            if not nombre:
                continue

            tel = parsear_telefono(_cell(idx_telefono))
            if tel is None:
                continue  # skip silencioso

            checkin_date = parsear_fecha(_cell(idx_checkin))

            # Extraer IDs de tipología {12345} y cruzar con maestro
            tip_val = _cell(idx_tipologia)
            apartamentos: list[str] = []
            if tip_val is not None:
                ids = _RE_IDS_TIP.findall(str(tip_val))
                for id_tip in ids:
                    apt = apt_repo.get_by_id_externo(empresa_id, id_tip)
                    if apt:
                        apartamentos.append(apt.nombre)

            registros.append({
                "nombre": nombre,
                "telefono": tel,
                "checkin_date": checkin_date,
                "apartamentos": apartamentos,
            })

        logger.info(
            "XLSX contactos: %d filas con teléfono procesadas",
            len(registros),
        )
        return registros, errors

    finally:
        wb.close()


def _build_contact_payload(contacto: ContactoAgrupado, prefs: dict) -> dict:
    """Construye el payload de Google People API para un ContactoAgrupado."""
    nombre_compuesto = aplicar_plantilla(
        plantilla=prefs.get("plantilla", "{FECHA} - {APT} - {NOMBRE}"),
        fecha=contacto.checkin_date,
        apartamentos=contacto.apartamentos,
        nombre=contacto.nombre,
        formato_fecha=prefs.get("formato_fecha_salida", "YYMMDD"),
        separador_apt=prefs.get("separador_apt", ", "),
    )
    return {
        "names": [{"givenName": nombre_compuesto}],
        "phoneNumbers": [{"value": contacto.telefono}],
    }


def _log_sync(empresa_id: str, estado: str, num: int, detalle: str) -> None:
    from app.h_maestro_apartamentos.repository import create_sync_log
    create_sync_log(
        empresa_id=empresa_id,
        origen=ORIGEN_GOOGLE_CONTACTS,
        estado=estado,
        num_registros=num,
        detalle=detalle,
    )


def _ultimo_sync(empresa_id: str) -> str | None:
    from app.h_sincronizador_contactos.model import LogSincronizacion
    log = (
        LogSincronizacion.query
        .filter_by(empresa_id=empresa_id, origen=ORIGEN_GOOGLE_CONTACTS)
        .order_by(LogSincronizacion.created_at.desc())
        .first()
    )
    return log.created_at.isoformat() if log else None


def _flatten(messages: dict) -> list[str]:
    errors: list[str] = []
    for field, msgs in messages.items():
        if isinstance(msgs, list):
            for msg in msgs:
                errors.append(f"{field}: {msg}")
        elif isinstance(msgs, dict):
            for sub in msgs.values():
                if isinstance(sub, list):
                    errors.extend(f"{field}: {m}" for m in sub)
    return errors
