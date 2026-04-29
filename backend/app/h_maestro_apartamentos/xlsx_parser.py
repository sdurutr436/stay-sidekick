"""Parser de archivos Excel (.xlsx) para importación de apartamentos.

Soporta dos modos de detección de columnas:

1. **Por cabecera** (por defecto): busca columnas por nombre en la primera fila.
2. **Por número de columna** (cuando se pasa `col_override`): usa los números de
   columna configurados por la empresa, ignorando las cabeceras. Los números son
   1-indexados (1 = columna A). Un valor de 0 significa "no configurado".

Columnas reconocidas en modo cabecera (al menos 'id' y 'nombre' son obligatorias):
- id / id_externo / identificador       → id_externo
- nombre / name / propiedad / apartment  → nombre
- direccion / dirección / address        → direccion
- ciudad / city                          → ciudad
"""

import logging
from dataclasses import dataclass
from io import BytesIO

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

_HEADER_MAP: dict[str, str] = {
    "id": "id_externo",
    "id_externo": "id_externo",
    "identificador": "id_externo",
    "nombre": "nombre",
    "name": "nombre",
    "propiedad": "nombre",
    "apartment": "nombre",
    "direccion": "direccion",
    "dirección": "direccion",
    "address": "direccion",
    "ciudad": "ciudad",
    "city": "ciudad",
}


@dataclass
class XlsxApartment:
    """Fila normalizada de un archivo Excel."""
    id_pms: str | None      # siempre columna 1; None si la celda está vacía
    id_externo: str         # código de tipología (columna configurada o detectada)
    nombre: str
    direccion: str | None
    ciudad: str | None


def parse_xlsx(
    file_bytes: bytes,
    col_override: dict | None = None,
) -> tuple[list[XlsxApartment], list[str]]:
    """Parsea un archivo .xlsx y devuelve apartamentos normalizados.

    Parameters
    ----------
    file_bytes:
        Contenido binario del archivo .xlsx.
    col_override:
        Diccionario con números de columna 1-indexados configurados por la empresa.
        Claves: col_id_externo, col_nombre, col_direccion, col_ciudad.
        Un valor de 0 significa "no configurado para este campo".
        Si None o todos los campos requeridos son 0, se usa detección por cabecera.

    Returns
    -------
    tuple[list[XlsxApartment], list[str]]
        (apartamentos, errores). Si hay errores críticos, la lista de
        apartamentos estará vacía.
    """
    errors: list[str] = []

    try:
        wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        return [], [f"No se pudo abrir el archivo Excel: {exc}"]

    try:
        ws = wb.active
        if ws is None:
            return [], ["El archivo Excel no tiene hojas de cálculo."]

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return [], ["El archivo Excel debe tener cabeceras + al menos una fila de datos."]

        # ── Construir mapa de columnas ──────────────────────────────────────
        use_override = _should_use_override(col_override)

        if use_override:
            col_map = _build_col_map_from_override(col_override)  # type: ignore[arg-type]
            data_rows = rows[1:]   # fila 0 son cabeceras, se omite igualmente
            row_offset = 2
        else:
            col_map, header_errors = _build_col_map_from_headers(rows[0])
            errors.extend(header_errors)
            if errors:
                return [], errors
            data_rows = rows[1:]
            row_offset = 2

        # ── Verificar campos obligatorios ───────────────────────────────────
        found_fields = set(col_map.values())
        missing = []
        if "id_externo" not in found_fields:
            missing.append("ID externo (id_externo / identificador / Código)")
        if "nombre" not in found_fields:
            missing.append("Nombre (nombre / name / propiedad)")
        if missing:
            return [], [f"Columna no encontrada: {m}" for m in missing]

        # ── Procesar filas ──────────────────────────────────────────────────
        result: list[XlsxApartment] = []
        for row_idx, row in enumerate(data_rows, start=row_offset):
            # id_pms: SIEMPRE columna 1 (índice 0), independientemente de la config
            id_pms_val: str | None = None
            if len(row) > 0 and row[0] is not None:
                raw = str(row[0]).strip()
                if raw:
                    id_pms_val = raw

            record: dict[str, str | None] = {
                "id_externo": None,
                "nombre": None,
                "direccion": None,
                "ciudad": None,
            }

            for col_idx, field in col_map.items():
                if col_idx < len(row) and row[col_idx] is not None:
                    record[field] = str(row[col_idx]).strip()

            if not record["id_externo"]:
                errors.append(f"Fila {row_idx}: falta el ID externo del apartamento.")
                continue
            if not record["nombre"]:
                errors.append(f"Fila {row_idx}: falta el nombre del apartamento.")
                continue

            result.append(XlsxApartment(
                id_pms=id_pms_val,
                id_externo=record["id_externo"],
                nombre=record["nombre"],
                direccion=record["direccion"],
                ciudad=record["ciudad"],
            ))

        logger.info("XLSX: %d filas parseadas, %d errores", len(result), len(errors))
        return result, errors

    finally:
        wb.close()


def _should_use_override(col_override: dict | None) -> bool:
    if not col_override:
        return False
    return (
        (col_override.get("col_id_externo") or 0) > 0
        and (col_override.get("col_nombre") or 0) > 0
    )


def _build_col_map_from_override(col_override: dict) -> dict[int, str]:
    """Construye col_map desde números de columna 1-indexados."""
    mapping = {
        "col_id_externo": "id_externo",
        "col_nombre": "nombre",
        "col_direccion": "direccion",
        "col_ciudad": "ciudad",
    }
    col_map: dict[int, str] = {}
    for key, field in mapping.items():
        num = col_override.get(key) or 0
        if num > 0:
            col_map[num - 1] = field  # 1-indexed → 0-indexed
    return col_map


def _build_col_map_from_headers(header_row: tuple) -> tuple[dict[int, str], list[str]]:
    """Construye col_map buscando cabeceras por nombre."""
    col_map: dict[int, str] = {}
    for i, header in enumerate(header_row):
        if header is None:
            continue
        key = str(header).strip().lower()
        field = _HEADER_MAP.get(key)
        if field:
            col_map[i] = field

    errors: list[str] = []
    found = set(col_map.values())
    if "id_externo" not in found:
        errors.append("No se encontró columna de ID (id, id_externo, identificador).")
    if "nombre" not in found:
        errors.append("No se encontró columna de nombre (nombre, name, propiedad, apartment).")

    return col_map, errors
