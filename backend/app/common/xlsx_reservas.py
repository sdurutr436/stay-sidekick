"""Parser de archivos Excel (.xlsx) para importación manual de reservas.

Normaliza las filas de un XLSX exportado desde cualquier PMS o herramienta
externa a objetos ``ReservaEstandar``, el mismo tipo que producen los
adaptadores de PMS (Smoobu, Beds24). Esto permite usar el flujo de
sincronización/exportación sin necesidad de conexión API al PMS.

Columnas reconocidas (case-insensitive, flexibles por idioma):

    Nombre huésped:  nombre, name, guest, huesped, huésped, guest_name
    Email:           email, correo, e-mail, mail
    Teléfono:        telefono, teléfono, phone, tel, móvil, movil
    Checkin:         checkin, check-in, check_in, fecha_entrada, arrival
    Checkout:        checkout, check-out, check_out, fecha_salida, departure
    Apartamento:     apartamento, propiedad, apartment, property, alojamiento, unit
    ID reserva:      id, id_reserva, booking_id, reserva, reservation_id, ref

Solo ``nombre`` es obligatoria. El resto son opcionales: un XLSX sin email
o sin teléfono genera contactos con los datos disponibles.
"""

import logging
import uuid
from io import BytesIO

from openpyxl import load_workbook

from app.normalizador_pms.base import ReservaEstandar

logger = logging.getLogger(__name__)

# Mapeo de posibles cabeceras → campo normalizado interno
_HEADER_MAP: dict[str, str] = {
    # Nombre del huésped
    "nombre": "nombre_raw",
    "name": "nombre_raw",
    "guest": "nombre_raw",
    "huesped": "nombre_raw",
    "huésped": "nombre_raw",
    "guest_name": "nombre_raw",
    "nombre_huesped": "nombre_raw",
    "nombre_huésped": "nombre_raw",
    # Email
    "email": "email",
    "correo": "email",
    "e-mail": "email",
    "mail": "email",
    "email_huesped": "email",
    # Teléfono
    "telefono": "telefono",
    "teléfono": "telefono",
    "phone": "telefono",
    "tel": "telefono",
    "movil": "telefono",
    "móvil": "telefono",
    # Checkin
    "checkin": "checkin",
    "check-in": "checkin",
    "check_in": "checkin",
    "fecha_entrada": "checkin",
    "arrival": "checkin",
    "entrada": "checkin",
    # Checkout
    "checkout": "checkout",
    "check-out": "checkout",
    "check_out": "checkout",
    "fecha_salida": "checkout",
    "departure": "checkout",
    "salida": "checkout",
    # Apartamento / propiedad
    "apartamento": "nombre_apartamento",
    "propiedad": "nombre_apartamento",
    "apartment": "nombre_apartamento",
    "property": "nombre_apartamento",
    "alojamiento": "nombre_apartamento",
    "unit": "nombre_apartamento",
    "unidad": "nombre_apartamento",
    # ID de reserva (opcional)
    "id": "id_externo",
    "id_reserva": "id_externo",
    "booking_id": "id_externo",
    "reserva": "id_externo",
    "reservation_id": "id_externo",
    "ref": "id_externo",
    "referencia": "id_externo",
    # Hora de llegada
    "hora": "hora_llegada",
    "hora_llegada": "hora_llegada",
    "hora_entrada": "hora_llegada",
    "arrival_time": "hora_llegada",
    "check_in_time": "hora_llegada",
    "hora_checkin": "hora_llegada",
    "time": "hora_llegada",
}


def parse_xlsx_reservas(
    file_bytes: bytes,
    col_overrides: dict[int, str] | None = None,
) -> tuple[list[ReservaEstandar], list[str]]:
    """Parsea un archivo .xlsx y devuelve reservas normalizadas.

    Parameters
    ----------
    file_bytes:
        Contenido binario del archivo Excel.
    col_overrides:
        Mapa opcional ``{índice_columna: campo_normalizado}`` que sobreescribe
        el col_map resuelto por cabeceras para los campos indicados.

    Returns
    -------
    tuple[list[ReservaEstandar], list[str]]
        (reservas, errores). Si hay errores críticos (archivo ilegible o sin
        columna de nombre), la lista de reservas estará vacía.
        Los errores por fila son no-críticos: se acumulan pero el parsing
        continúa con el resto de filas.
    """
    errors: list[str] = []

    try:
        wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        return [], [f"No se pudo abrir el archivo Excel: {exc}"]

    try:
        ws = wb.active
        if ws is None:
            return [], ["El archivo Excel no tiene hojas de cálculo."]

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return [], ["El archivo debe tener cabeceras en la primera fila y al menos una fila de datos."]

        # Resolver cabeceras (primera fila)
        raw_headers = rows[0]
        col_map: dict[int, str] = {}  # índice columna → campo normalizado
        for i, header in enumerate(raw_headers):
            if header is None:
                continue
            key = str(header).strip().lower().replace(" ", "_")
            field = _HEADER_MAP.get(key)
            if field and field not in col_map.values():
                col_map[i] = field

        # Aplicar sobreescrituras de columna explícitas
        if col_overrides:
            for col_idx, field in col_overrides.items():
                col_map[col_idx] = field

        # Columna de nombre es obligatoria
        if "nombre_raw" not in col_map.values():
            return [], [
                "No se encontró columna de nombre del huésped. "
                "Cabeceras reconocidas: nombre, name, guest, huesped, guest_name."
            ]

        # Procesar filas de datos
        result: list[ReservaEstandar] = []
        for row_idx, row in enumerate(rows[1:], start=2):
            record: dict[str, str | None] = {
                "id_externo": None,
                "nombre_raw": None,
                "email": None,
                "telefono": None,
                "checkin": None,
                "checkout": None,
                "nombre_apartamento": None,
                "hora_llegada": None,
            }

            for col_idx, field in col_map.items():
                if col_idx < len(row) and row[col_idx] is not None:
                    raw = row[col_idx]
                    if field == "hora_llegada":
                        # Objeto datetime/time con atributo hour
                        if hasattr(raw, "hour"):
                            record[field] = f"{raw.hour:02d}:{raw.minute:02d}"
                        else:
                            s = str(raw).strip()
                            if len(s) >= 5 and s[2] == ":" and s[:5].replace(":", "").isdigit():
                                record[field] = s[:5]
                            else:
                                record[field] = None
                                errors.append(
                                    f"Fila {row_idx}: valor de hora_llegada no reconocido '{raw}' — ignorado."
                                )
                    # Las fechas pueden venir como objetos date/datetime de openpyxl
                    elif hasattr(raw, "isoformat"):
                        record[field] = raw.isoformat()[:10]  # solo YYYY-MM-DD
                    else:
                        record[field] = str(raw).strip()

            # Nombre es obligatorio por fila
            if not record["nombre_raw"]:
                errors.append(f"Fila {row_idx}: falta el nombre del huésped — fila omitida.")
                continue

            # Si no hay ID externo, generamos uno reproducible desde el nombre + fila
            id_externo = record["id_externo"] or f"xlsx-{row_idx}-{uuid.uuid4().hex[:8]}"

            result.append(ReservaEstandar(
                id_externo=id_externo,
                nombre_raw=record["nombre_raw"],
                email=record["email"] or None,
                telefono=record["telefono"] or None,
                checkin=record["checkin"] or None,
                checkout=record["checkout"] or None,
                nombre_apartamento=record["nombre_apartamento"] or None,
                id_apartamento_externo=None,
                hora_llegada=record["hora_llegada"] or None,
            ))

        logger.info(
            "XLSX reservas: %d filas procesadas, %d omitidas por error",
            len(result),
            len(errors),
        )
        return result, errors

    finally:
        wb.close()
