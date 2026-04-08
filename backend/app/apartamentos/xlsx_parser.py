"""Parser de archivos Excel (.xlsx) para importación de apartamentos.

El archivo debe tener cabeceras en la primera fila. El parser busca
columnas por nombre (case-insensitive), no por posición, para tolerar
distintos formatos de exportación.

Columnas reconocidas (al menos 'id' y 'nombre' son obligatorias):
- id / id_externo / identificador       → id_externo
- nombre / name / propiedad / apartment  → nombre
- direccion / dirección / address        → direccion
- ciudad / city                          → ciudad
"""

import logging
from dataclasses import dataclass
from io import BytesIO

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Mapeo de posibles nombres de cabecera → campo normalizado
_HEADER_MAP: dict[str, str] = {
    "id": "id_externo",
    "id_externo": "id_externo",
    "identificador": "id_externo",
    "nombre": "nombre",
    "name": "nombre",
    "propiedad": "nombre",
    "apartment": "nombre",
    "direccion": "direccion",
    "dirección": "direccion",
    "address": "direccion",
    "ciudad": "ciudad",
    "city": "ciudad",
}


@dataclass
class XlsxApartment:
    """Fila normalizada de un archivo Excel."""
    id_externo: str
    nombre: str
    direccion: str | None
    ciudad: str | None


def parse_xlsx(file_bytes: bytes) -> tuple[list[XlsxApartment], list[str]]:
    """Parsea un archivo .xlsx y devuelve apartamentos normalizados.

    Returns
    -------
    tuple[list[XlsxApartment], list[str]]
        (apartamentos, errores). Si hay errores críticos, la lista
        de apartamentos estará vacía.
    """
    errors: list[str] = []

    try:
        wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        return [], [f"No se pudo abrir el archivo Excel: {exc}"]

    try:
        ws = wb.active
        if ws is None:
            return [], ["El archivo Excel no tiene hojas de cálculo."]

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return [], ["El archivo Excel debe tener cabeceras + al menos una fila de datos."]

        # Resolver cabeceras
        raw_headers = rows[0]
        col_map: dict[int, str] = {}  # índice de columna → campo normalizado
        for i, header in enumerate(raw_headers):
            if header is None:
                continue
            key = str(header).strip().lower()
            field = _HEADER_MAP.get(key)
            if field:
                col_map[i] = field

        # Verificar columnas obligatorias
        found_fields = set(col_map.values())
        if "id_externo" not in found_fields:
            errors.append("No se encontró columna de ID (id, id_externo, identificador).")
        if "nombre" not in found_fields:
            errors.append("No se encontró columna de nombre (nombre, name, propiedad, apartment).")
        if errors:
            return [], errors

        # Procesar filas
        result: list[XlsxApartment] = []
        for row_idx, row in enumerate(rows[1:], start=2):
            record: dict[str, str | None] = {
                "id_externo": None,
                "nombre": None,
                "direccion": None,
                "ciudad": None,
            }

            for col_idx, field in col_map.items():
                if col_idx < len(row) and row[col_idx] is not None:
                    record[field] = str(row[col_idx]).strip()

            # Validar campos obligatorios por fila
            if not record["id_externo"]:
                errors.append(f"Fila {row_idx}: falta el ID del apartamento.")
                continue
            if not record["nombre"]:
                errors.append(f"Fila {row_idx}: falta el nombre del apartamento.")
                continue

            result.append(XlsxApartment(
                id_externo=record["id_externo"],
                nombre=record["nombre"],
                direccion=record["direccion"],
                ciudad=record["ciudad"],
            ))

        logger.info("XLSX: %d filas parseadas, %d errores", len(result), len(errors))
        return result, errors
    finally:
        wb.close()
